import openpyxl as opyxl
import os

script_dir = os.path.dirname(os.path.realpath(__file__))

if __name__ == '__main__':
	wb_in = opyxl.load_workbook(script_dir + "/pll-reconfiguration-calculator.xlsx")
	wb_out = opyxl.Workbook()
	wb_out_cal = wb_out.create_sheet(title = 'Cal')
	wb_out_cal = wb_out['Cal']
	wb_in_cal = wb_in['Cal']
	for j in range(1, 1000):
		for i in list(map(chr, range(65, 90))):
			wb_out_cal[i+str(j)].value = wb_in_cal[i+str(j)].value
		for i in list(map(chr, range(65, 90))):
			wb_out_cal["A"+i+str(j)].value = wb_in_cal["A"+i+str(j)].value
		for i in list(map(chr, range(65, 90))):
			wb_out_cal["B"+i+str(j)].value = wb_in_cal["A"+i+str(j)].value
	# Copy Sheet2
	wb_out_sheet2 = wb_out.create_sheet(title = 'Sheet2')
	wb_out_sheet2 = wb_out['Sheet2']
	wb_in_sheet2 = wb_in['Sheet2']
	for j in range(1, 1000):
		for i in list(map(chr, range(65, 90))):
			wb_out_sheet2[i+str(j)].value = wb_in_sheet2[i+str(j)].value
		for i in list(map(chr, range(65, 90))):
			wb_out_sheet2["A"+i+str(j)].value = wb_in_sheet2["A"+i+str(j)].value
		for i in list(map(chr, range(65, 90))):
			wb_out_sheet2["B"+i+str(j)].value = wb_in_sheet2["A"+i+str(j)].value
	wb_out.save(filename='hacked_pll_config.xlsx')